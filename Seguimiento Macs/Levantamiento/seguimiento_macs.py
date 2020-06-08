from openpyxl.styles import Font, PatternFill, Border, Alignment, Side, NamedStyle
from script_levantamiento import recopila, ping
from script_procesamiento import procesa
from joblib import Parallel, delayed
import multiprocessing
import subprocess
import openpyxl
import glob
import re



def get_isla_ips(equipos_isla):
	'''	
		Ciclo para obtener las ips de los equipos de la Isla a partir del archivo de Excel informacion_macs
		Empieza en la fila 2 columna 1 de la hoja 'Equipos Isla'
		Retorna una lista de strings
	'''
	lista_equipos=[]
	row, column = 2, 1
	ip_equipo = equipos_isla.cell(row=row,column=column).value
	while ip_equipo != None:
		lista_equipos.append(ip_equipo)
		row+=1
		ip_equipo = equipos_isla.cell(row=row,column=column).value
	return lista_equipos

def write_errors_to_excel(equipos_isla):
	'''
		Escribe errores de conexion y de ping ocurridos durante el levantamiento de informacion
	'''
	lista_equipos = glob.glob('C:/Users/Ileam Montoya/Dropbox/PYTHON/Python Scripts/CLARO/Seguimiento Macs/Levantamiento/data/*.txt')
	error_dict={}

	for resultado in lista_equipos:
		if 'ERROR' in resultado:
			error_ip=re.findall('.+\\\(.+) - (\d+\.\d+\.\d+\.\d+)',resultado)
			error_dict[error_ip[0][1]] = error_ip[0][0]
	row, column = 2, 1
	ip_equipo = equipos_isla.cell(row=row,column=column).value
	while ip_equipo != None:
		if ip_equipo in error_dict:
			equipos_isla.cell(row=row,column=3).value = error_dict[ip_equipo]
		row+=1
		ip_equipo = equipos_isla.cell(row=row,column=column).value

def get_isla_macs(macs_isla):
	'''	
		Ciclo para obtener direcciones mac a partir del archivo de Excel informacion_macs
		Empieza en la fila 2 columna 2 de la hoja 'MACS'
		Retorna una lista de strings
	'''
	lista_macs=[]
	row, column = 2, 2
	mac = macs_isla.cell(row=row,column=column).value
	while mac != None:
		ping(macs_isla.cell(row=row,column=1).value)
		lista_macs.append(mac)
		row+=1
		mac = macs_isla.cell(row=row,column=column).value
	return lista_macs	

	
	
if __name__ == "__main__":

	#Carga archivo excel y las pestañas necesarias
	path_excel = 'C:/Users/Ileam Montoya/Dropbox/PYTHON/Python Scripts/CLARO/Seguimiento Macs/Levantamiento/informacion_macs.xlsx'
	informacion_macs = openpyxl.load_workbook(path_excel)
	equipos_isla = informacion_macs['Equipos Isla']
	macs_isla = informacion_macs['MACS']

	#Genera listas de equipos en la isla y macs a procesar
	#En caso de querer procesar solo algunos equipos y no la lista completa, colocar los equipos en la lista comentada y activar la linea
	lista_equipos = get_isla_ips(equipos_isla)
	# lista_equipos = ['10.179.28.23']
	lista_macs = get_isla_macs(macs_isla)

	#Parte que multiprocesa el levantmiento de las tablas mac y vecinos de cada equipo
	#La funcion 'recopila' se encuentra en el archivo de python script_levantamiento.py
	num_cores = multiprocessing.cpu_count() - 2
	results = Parallel(n_jobs=num_cores)(delayed(recopila)(ip) for ip in lista_equipos)

	#Parte que lee los archivos de texto obtenidos durante el levantamiento de informacion y los procesa para escribir los resultados en el excel
	#La funcion 'procesa' se encuentra en el archivo de python script_procesamiento.py
	procesa(lista_macs,macs_isla)

	#Escribe los errores ocurridos en los equipos durante el levantamiento de informacion
	write_errors_to_excel(equipos_isla)

	#Guarda el excel
	informacion_macs.save(path_excel)
