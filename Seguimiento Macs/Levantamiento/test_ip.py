import openpyxl
import subprocess

def ping(host):
	command = 'ping -n 1 {}'.format(host)
	return subprocess.call(command, stdout=subprocess.DEVNULL) == 0

def test_ping(equipos_isla):
	'''	
		Ciclo para obtener las ips de los equipos de la Isla a partir del archivo de Excel informacion_macs
		Empieza en la fila 2 columna 1 de la hoja 'Equipos Isla'
		Retorna una lista de strings
	'''
	row, column = 2, 1
	ip = equipos_isla.cell(row=row,column=column).value
	while ip != None:
		if ping(ip):
			print('Ping exitoso a IP {}'.format(ip))
			equipos_isla.cell(row=row,column=3).value = 'YES'
		else:
			print('Ping fallido a IP {}'.format(ip))
			equipos_isla.cell(row=row,column=3).value = 'NO'
		row+=1
		ip = equipos_isla.cell(row=row,column=column).value
	
if __name__ == "__main__":

	#Carga archivo excel y las pestañas necesarias
	path_excel = 'C:/Users/Ileam Montoya/Dropbox/PYTHON/Python Scripts/CLARO/Seguimiento Macs/Levantamiento/informacion_macs.xlsx'
	informacion_macs = openpyxl.load_workbook(path_excel)
	equipos_isla = informacion_macs['testing']

	test_ping(equipos_isla)
	informacion_macs.save(path_excel)

