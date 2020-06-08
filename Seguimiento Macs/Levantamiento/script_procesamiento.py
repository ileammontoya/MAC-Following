from seguimiento_macs import get_isla_macs
import openpyxl
import glob
import re


def procesa(macs,macs_isla):
	'''
		Funcion que procesa los archivos de texto generados por script_levantamiento.py para obtener el camino seguido para cada mac de la lista
		Resultados se escriben directamente sobre el archivo de excel de donde se levanta la informacion
	'''
	
	for row, mac in enumerate(macs,2):
		column = 3
		print('\n\nProcessing mac {}'.format(mac))
		primary_pe = 'GNCYGTLON2D1D02A02EIM2' # PE Origen 
		secondary_pe ='GNCYGTLON2D1D01A02EIM3' # PE Secundario
		#Se comienza la busqueda en el archivo de texto de los PE
		seed=primary_pe
		
		lista_equipos = glob.glob('C:/Users/Ileam Montoya/Dropbox/PYTHON/Python Scripts/CLARO/Seguimiento Macs/Levantamiento/data/*.txt')
		dict_equipos={}
		next = True
		#Ciclo para saltar a cada equipo a donde se siga la mac
		while next:
			file, next_text, next_hop = '', '', ''
			print(seed)
			#Se escribe el equipo donde se esta buscando la mac en el archivo de excel
			macs_isla.cell(row=row,column=column).value = seed
			column+=1
			#Ciclo para leer la informacion del archivo de texto del equipo
			for equipo_ip in lista_equipos:
				if seed in equipo_ip:
					read_del=equipo_ip
					reading=open(read_del,'r')
					file = reading.read()
					break
			#si la variable file contiene datos, se continua el loop
			if file:
				next = True
			#si la variable file se encuentra vacia, se termina el loop
			else:
				next = False

			#A partir de la informacion del archivo se capturan las macs e interfaces de salida
			mac_capture=re.findall('(....\.....\.....)\t(.+)\n',file)
			if mac_capture:
				for mac_data, interface in mac_capture:
					if mac in mac_data:
						next_hop = interface
				#En caso de no capturar un proximo salto, se debe verificar el estado de la mac en ese equipo
				if not next_hop:
					seed = 'VERIFICAR SALTO'
			#A partir de la informacion del archivo se capturan las interfaces de vecindad y proximos equipos	
			neighbor_capture=re.findall('Interface (.+) +- Neighbor (.+)\n',file)
			if neighbor_capture:
				for interface, neighbor in neighbor_capture:
					#si el proximo salto de la mac coincide con la interfaz de vecindad, se escribe esa interfaz en el archivo excel
					#Tambien se cambia el seed al proximo equipo
					if next_hop == interface:
						print(interface)
						macs_isla.cell(row=row,column=column).value = interface
						column+=1
						seed = neighbor


if __name__ == "__main__":

	#Carga archivo excel y las pestañas necesarias
	path_excel = 'C:/Users/Ileam Montoya/Dropbox/PYTHON/Python Scripts/CLARO/Seguimiento Macs/Levantamiento/informacion_macs.xlsx'
	informacion_macs = openpyxl.load_workbook(path_excel)
	macs_isla = informacion_macs['MACS']
	macs = get_isla_macs(macs_isla)

	# macs= ['2c97.b1e8.4de6','8038.bc1d.d7be','4c1f.cc2f.1e74','4482.e5d2.1866','9c74.1a6f.04ec']

	procesa(macs,macs_isla)